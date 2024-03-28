import os

os.environ["CUBLAS_WORKSPACE_CONFIG"] = ":4096:8"
import torch
import argparse
import torch.utils.data.distributed
import numpy as np

from classification.models import get_model
from core.utils.config import (
    config,
    load_transfer_config,
    update_config_from_wandb,
)
from core.utils.logging import logger

import yaml
from easydict import EasyDict
from itertools import product


def get_parser():
    parser = argparse.ArgumentParser(
        formatter_class=argparse.ArgumentDefaultsHelpFormatter
    )

    parser.add_argument(
        "-w", "--wandb_sweep_config",
        nargs='+',  # Accept one or more values
        type=str,
        default="./policies/test_policy.yaml",
        help="Which sweep config file to call for training",
    )

    parser.add_argument(
        "-o", "--output_file",
        type=str,
        default="output.xlsx",
        help="Specify the output file store the results of best models. Supported formats are: .xlsx,.xlsm,.xltx,.xltm",
    )

    parser.add_argument(
        "--config-file",
        type=str,
        default="transfer.yaml",
        help="Which config file to call for training",
    )

    parsed = parser.parse_args()


    return parsed.wandb_sweep_config, parsed.config_file, parsed.output_file


def generate_config(file_path):
    """
    Generate all possible configurations from a Weights & Biases (wandb) sweep config file.

    Parameters:
        file_path (str): The path to the wandb sweep config file.
                         For example: 'test_policy.yaml'.
                         
    Returns:
        list: A list of dictionaries, each containing a possible configuration.
    """
    # generate các config khả thi từ file sweep
    def _read_config(file_path):
        with open(file_path, 'r') as file:
            config = yaml.safe_load(file)

        parameters = config.get('parameters', {})
        parameters_dict = {}

        for key, value in parameters.items():
            values = value.get('values', [])
            if len(values) == 1:
                parameters_dict[key] = values[0]
            else:
                parameters_dict[key] = values

        return EasyDict(parameters_dict)
    wandb_config = _read_config(file_path)

    multi_value_keys = {key: wandb_config[key] for key in wandb_config if isinstance(wandb_config[key], list)}
    combinations_values = product(*multi_value_keys.values())

    combinations = []
    for values in combinations_values:
        combination = wandb_config.copy()
        for key, value in zip(multi_value_keys.keys(), values):
            combination[key] = value
        combinations.append(EasyDict(combination))

    return combinations
    


def load_best_model(model):
    model_fname = os.path.join(os.path.join(config.run_dir, "checkpoint"), "ckpt.best.pth")
    checkpoint = None
    if os.path.exists(model_fname):
        checkpoint = torch.load(model_fname, map_location="cpu")
    else:
        print("[Warning] No best model is found at ", model_fname, " !!")
    return checkpoint

def filter_config(wandb_configs, filter_key):
    """
    Filter configurations based on a specified key.

    Parameters:
        wandb_configs (list): A list of configurations generated from the sweep file.
        filter_key (str): The key used for filtering configurations.

    Returns:
        dict: A dictionary where keys are values of the filter_key and values are lists of configurations
              containing the corresponding key value.

    Example:
        If there are 4 configurations from the sweep file {value1: A, B; value2: X, Y}:
        - {value1: A; value2: X}, {value1: A; value2: Y}, {value1: B; value2: X}, and {value1: B; value2: Y}.
        If filtered with filter_key as 'value1', the output will be:
        {
            "A": [{value1: A; value2: X}, {value1: A; value2: Y}],
            "B": [{value1: B; value2: X}, {value1: B; value2: Y}]
        }
    """
    classified_dicts = {}

    # Iterate through each easydict in the array
    for easy_dict in wandb_configs:
        # Get the value of the key to filter
        key = easy_dict.get(filter_key)
        # Check if the key has been added to the dictionary
        if key not in classified_dicts:
            # If not, add a new key-value pair with the key and value as a list containing the current easydict
            classified_dicts[key] = [easy_dict]
        else:
            # If it already exists, append the current easydict to the list corresponding to the key
            classified_dicts[key].append(easy_dict)
    return classified_dicts


def filter_config_recursive(wandb_configs, filter_keys):
    """
    Recursively filter configurations based on specified keys.

    Parameters:
        wandb_configs (list): A list of configurations generated from the sweep file.
        filter_keys (list): A list containing keys that the function will filter by.

    Returns:
        dict: A dictionary where keys are values of the filter_keys and values are dictionaries
              containing configurations with corresponding key values.

    Example:
        If filter_keys = ['value1', 'value2'] and there are 4 configurations from the sweep file
        {value1: A, B; value2: X, Y, value3: G}:
        - {value1: A; value2: X, value3: G}, {value1: A; value2: Y, value3: G},
          {value1: B; value2: X, value3: G}, and {value1: B; value2: Y, value3: G}.
        If filtered with filter_keys as ['value1', 'value2'], the output will be:
        {
            "A": {
                "X": {value1: A; value2: X, value3: G},
                "Y": {value1: A; value2: Y, value3: G}
            },
            "B": {
                "X": {value1: B; value2: X, value3: G},
                "Y": {value1: B; value2: Y, value3: G}
            }
        }
    """
    # If filter_keys is empty, return wandb_configs
    if not filter_keys:
        return wandb_configs
    
    filtered_configs = {}
    # Iterate through the elements of filter_keys[0]
    for key, cof in filter_config(wandb_configs, filter_keys[0]).items():
        # Use recursion to handle the remaining keys in filter_keys
        filtered_configs[key] = filter_config_recursive(cof, filter_keys[1:])
    
    return filtered_configs


def calculate_mean_and_deviation(data):
    """
    Calculate mean and standard deviation of the given data.

    Parameters:
        data (list): A list of numeric values.

    Returns:
        tuple: A tuple containing the mean and standard deviation of the data.

    Example:
        If data = [1, 2, 3, 4, 5], the calculated mean and standard deviation will be:
        (3.0, 1.58)
    """
    # Calculate the mean
    mean = np.mean(data)
    
    squared_diff = sum((x - mean) ** 2 for x in data)
    # Calculate the variance
    variance = squared_diff / (len(data) - 1)

    # Calculate the standard deviation
    deviation = np.sqrt(variance)
    
    # Round the results to 2 decimal places
    mean = round(mean, 2)
    deviation = round(deviation, 2)

    return mean, deviation
    

from openpyxl import load_workbook, Workbook

def export_to_excel(data, filename):
    """
    Export data to an Excel file.

    Parameters:
        data (list): A list of lists (element-list) containing data to be exported.
                     Each element-list should represent a row in the Excel file.
        filename (str): The filename/path of the Excel file to export the data to.

    Returns:
        None

    Description:
        This function exports the given data to an Excel file specified by the filename.
        If the file already exists, it loads the existing workbook and updates the data.
        If the file doesn't exist, it creates a new workbook and adds the data to it.
        The data should be in the format where each tuple represents a row in the Excel file.
        The function handles updating existing rows and adding new rows as necessary.
    """
    try:
        # Load workbook
        wb = load_workbook(filename)
        # Load the active sheet
        sheet = wb.active
        # Get column indices
        net_name_col = 1
        scheme_col = 2
        budget_col = 3
        initialization_col = 4
        neuron_selection_col = 5
        
        # Find existing rows
        existing_rows = {}
        for row in range(2, sheet.max_row + 1):
            net_name = sheet.cell(row=row, column=net_name_col).value
            scheme = sheet.cell(row=row, column=scheme_col).value
            budget = sheet.cell(row=row, column=budget_col).value
            initialization = sheet.cell(row=row, column=initialization_col).value
            neuron_selection = sheet.cell(row=row, column=neuron_selection_col).value
            existing_rows[(net_name, scheme, budget, initialization, neuron_selection)] = row
        
        # Write or overwrite data
        for row_data in data:
            net_name, scheme, budget, initialization, neuron_selection, valid_best, valid_best_epoch, test_top1_valid_best, best_epoch_mean = row_data
            key = (net_name, scheme, budget, initialization, neuron_selection)
            if key in existing_rows:
                row_index = existing_rows[key]
                # sheet.cell(row=row_index, column=3).value = budget
                sheet.cell(row=row_index, column=6).value = valid_best
                sheet.cell(row=row_index, column=7).value = valid_best_epoch
                sheet.cell(row=row_index, column=8).value = test_top1_valid_best
                sheet.cell(row=row_index, column=9).value = best_epoch_mean
            else:
                sheet.append(row_data)
        
        # Save the workbook
        wb.save(filename)
    except FileNotFoundError:
        # If file doesn't exist, create a new one
        wb = Workbook()
        ws = wb.active
        ws.append(["Model", "Scheme", "Budget", "Initialization", "Method", "valid/best", "valid/best at epoch", "test/top1 at valid/best", "Average Stopping Epoch"])
        for row_data in data:
            ws.append(row_data)
        wb.save(filename)


def iterate_configs(filtered_configs, filter_keys, excel_data_frame, current_level=0):
    """
    Recursively iterate through filtered configurations and log results to an Excel dataframe.

    Parameters:
        filtered_configs (dict): A dictionary containing filtered configurations.
                                 Keys are values of the filter_keys, and values are sub-dictionaries
                                 containing configurations with corresponding key values.
        filter_keys (list): A list containing keys used for filtering configurations.
        excel_data_frame (DataFrame): A Pandas DataFrame to log the results.
        current_level (int): Current level of recursion (default is 0).

    Returns:
        None

    Description:
        This function iterates through the filtered configurations recursively.
        It logs the results to the provided Excel DataFrame.
        The function logs mean and deviation results of 'val/best' and 'test/top1 at val/best'.
        It also logs the stopping epoch and budget information.

    Example:
        If filter_keys = ['value1', 'value2'] and there are configurations filtered as follows:
        
        {
            "A": {
                "X": {value1: A; value2: X, value3: G},
                "Y": {value1: A; value2: Y, value3: G}
            },
            "B": {
                "X": {value1: B; value2: X, value3: G},
                "Y": {value1: B; value2: Y, value3: G}
            }
        }

        The function will recursively log the result of run into 'excel_data_frame' in the following order:
        
        1st: It will log the result of run with configuration {value1: A; value2: X, value3: G}
        2nd: It will log the result of run with configuration {value1: A; value2: Y, value3: G}
        3rd: It will log the result of run with configuration {value1: B; value2: X, value3: G}
        4th: It will log the result of run with configuration {value1: B; value2: Y, value3: G}

    """
    if current_level == len(filter_keys):
        best_val = []
        test_top1_at_best_val = []
        best_at_epochs = []
        have_best_model = True
        # Iterate through each random seed
        for wandb_config in filtered_configs:
            if config.wandb_sweep:
                update_config_from_wandb(wandb_config)
            # Loading model
            if "mcunet" in config.net_config.net_name or config.net_config.net_name == "proxyless-w0.3":
                model, config.data_provider.image_size, description, total_neurons = get_model(config.net_config.net_name)
            else:
                model, total_neurons = get_model(config.net_config.net_name)
            # print("Seed: ", config.manual_seed)
            # Load best model
            best_checkpoint = load_best_model(model)
            # At each random seed
            if best_checkpoint == None:
                have_best_model = False
                # No value is logged
                test_top1_at_best_val.append(None)
                best_val.append(None)
                best_at_epochs.append(None)
                break
            test_top1_at_best_val.append(best_checkpoint["test_top1_at_best_val"])
            best_val.append(best_checkpoint["best_val"])
            best_at_epochs.append(best_checkpoint["epoch"])
        
        # Log mean and deviation result of 'val/best' and 'test/top1 at val/best'
        if have_best_model:
            test_top1_at_best_val_mean, test_top1_at_best_val_deviation = calculate_mean_and_deviation(test_top1_at_best_val)
            best_val_mean, best_val_deviation = calculate_mean_and_deviation(best_val)
        else:
            test_top1_at_best_val_mean, test_top1_at_best_val_deviation = None, None
            best_val_mean, best_val_deviation = None, None
        # Log stopping epoch:
        best_epoch_mean, best_epoch_deviation = calculate_mean_and_deviation(best_at_epochs)
        best_epoch_mean = str(best_epoch_mean) + "\u00B1" + str(best_epoch_deviation)
        # Log budget
        from general_utils import compute_update_budget
        if "fixed_budget" in wandb_config.scheme or "mcunet" in wandb_config.scheme:
            budget = config.NEq_config.budget
        else:
            budget = compute_update_budget(
                config.NEq_config.total_num_params, config.NEq_config.ratio
            )

        excel_data_frame.append([config.net_config.net_name,
                                wandb_config.scheme,
                                budget,
                                wandb_config.initialization,
                                wandb_config.neuron_selection,
                                str(best_val_mean) + "\u00B1" + str(best_val_deviation), 
                                str(best_at_epochs), 
                                str(test_top1_at_best_val_mean) + "\u00B1" + str(test_top1_at_best_val_deviation),
                                best_epoch_mean])
        print("[Result] Model ", config.net_config.net_name, " achieves best validation result at epoch ", best_at_epochs, " with test/top1 is ", test_top1_at_best_val_mean, "\u00B1", test_top1_at_best_val_deviation)
        print("-------------------------------------")
        return

    for item_key, sub_configs in filtered_configs.items():
        prefix = '-' * (len(filter_keys) - current_level)
        print(prefix, " With ", filter_keys[current_level], " is ", item_key, " ", prefix)
        # Recursion to navigate to the next level
        iterate_configs(sub_configs, filter_keys, excel_data_frame, current_level + 1)



def main(wandb_config_path, transfer_config_path, output_file):

    load_transfer_config(transfer_config_path)
    wandb_configs = generate_config(wandb_config_path)
    print("================ Model: ", wandb_configs[0]["net_name"], " ================ ")
    excel_data_frame = []
    
    """
    The program will filter combinations of configurations generated from the original sweep 
    file according to the order of keys in filter_keys (similar to how WanDB works)
    """
    # Here, random initialization is utilized for all cases except baseline, therefore it does not need to be included in filter keys
    if "fixed_budget" in wandb_config_path:
        """
        'filter_keys' does not contain 'manual_seed' because it will be manually defined
        in the iterate_configs() function. Results from different manual seeds will be averaged.
        """
        filter_keys = ["scheme", "budget", "neuron_selection"]
    else:
        # same
        filter_keys = ["scheme", "neuron_selection"]
    """
    From the configuration combinations, create a dictionary where the top-level key is the first
    element of 'filter_keys'; each value of these keys is a sub-dictionary; each sub-dictionary has
    keys as the next element of filter_keys ...; until at the lowest level, each value of each key
    is not a sub-dictionary, but a configuration
    """
    filtered_configs = filter_config_recursive(wandb_configs, filter_keys)

    # A recursive function to iterate through the layers of dictionaries created above, where 'manual_seed' is the deepest layer
    iterate_configs(filtered_configs, filter_keys, excel_data_frame)
    
    export_to_excel(data=excel_data_frame, filename=output_file)



if __name__ == "__main__":
    wandb_config_paths, transfer_config_path, output_file = get_parser()
    for wandb_config_path in wandb_config_paths:
        main(wandb_config_path, transfer_config_path, output_file)
